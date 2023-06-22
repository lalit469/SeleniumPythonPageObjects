import openpyxl


def getRowCount(path,sheetName):
    wb = openpyxl.load_workbook(path)
    ws = wb[sheetName]
    return ws.max_row

def getColCount(path,sheetName):
    wb = openpyxl.load_workbook(path)
    ws = wb[sheetName]
    return ws.max_column

def getCellValue(path,sheetName,row,column):
    wb = openpyxl.load_workbook(path)
    ws = wb[sheetName]
    return ws.cell(row=row,column=column).value

def setCellValue(path,sheetName,row,column,data):
    wb = openpyxl.load_workbook(path)
    ws = wb[sheetName]
    ws.cell(row=row,column=column).value=data
    wb.save(path)

path="..//excel//testData.xlsx"
sheetName="login"


rows=getRowCount(path,sheetName)
print(rows)

cols=getColCount(path,sheetName)
print(cols)

print(getCellValue(path,sheetName,4,2))

setCellValue(path,sheetName,15,2,"12345")


